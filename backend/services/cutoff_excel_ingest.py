"""
Excel ingestion for NEET UG 2025 cutoff rows.
"""

from __future__ import annotations

import re
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from sqlalchemy import delete, insert
from sqlalchemy.ext.asyncio import AsyncSession

from models.neet_ug_cutoff import NeetUg2025Cutoff


SKIP_SHEETS = {"COLLEGE MASTER"}


def _canon(text: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(text or "").strip().lower())


def _sheet_state_name(sheet_name: str) -> str:
    return "MCC" if sheet_name.strip().upper() == "MCC" else sheet_name.strip().title()


HEADER_ALIASES: Dict[str, str] = {
    "state": "state",
    "air": "air_rank",
    "airrank": "air_rank",
    "staterank": "state_rank",
    "code": "college_code",
    "collegecode": "college_code",
    "collegename": "college_name",
    "institutionname": "institution_name",
    "collegetype": "college_type",
    "course": "course",
    "category": "category",
    "subcategory": "sub_category",
    "seattype": "seat_type",
    "quota": "quota",
    "domicile": "domicile",
    "eligibility": "eligibility",
    "score": "score",
    "round": "round",
}


INT_FIELDS = {"air_rank", "state_rank", "score"}


def _to_int(value: object) -> Optional[int]:
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    try:
        return int(float(text))
    except Exception:
        return None


def _to_text(value: object, max_len: int) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return text[:max_len]


FIELD_MAX_LEN = {
    "state": 100,
    "college_code": 50,
    "college_name": 500,
    "institution_name": 500,
    "college_type": 100,
    "course": 100,
    "category": 100,
    "sub_category": 100,
    "seat_type": 100,
    "quota": 100,
    "domicile": 200,
    "eligibility": 500,
    "round": 50,
}


def _select_target_sheets(all_sheets: List[str], only_state: Optional[str]) -> List[str]:
    filtered = [s for s in all_sheets if s.strip().upper() not in SKIP_SHEETS]
    if not only_state:
        return filtered
    target = _canon(only_state)
    matched = [s for s in filtered if _canon(s) == target]
    if matched:
        return matched
    # Fallback: if workbook has one data sheet, allow mapping to requested state.
    return filtered if len(filtered) == 1 else []


def parse_cutoff_workbook(
    file_path: str,
    only_state: Optional[str] = None,
) -> Tuple[List[Dict], Dict[str, int], List[str]]:
    wb = load_workbook(file_path, data_only=True, read_only=True)
    target_sheets = _select_target_sheets(wb.sheetnames, only_state)
    if only_state and not target_sheets:
        raise ValueError(f"No sheet found for state '{only_state}'")

    all_rows: List[Dict] = []
    state_counts: Dict[str, int] = {}
    ingested_states: List[str] = []

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        headers = list(next(rows, []) or [])
        header_map: Dict[int, str] = {}
        for idx, header in enumerate(headers):
            mapped = HEADER_ALIASES.get(_canon(header))
            if mapped:
                header_map[idx] = mapped

        state_name = _sheet_state_name(only_state or sheet_name) if (only_state and len(target_sheets) == 1) else _sheet_state_name(sheet_name)
        count = 0

        for row in rows:
            if not row:
                continue
            payload: Dict[str, object] = {
                "state": state_name,
                "air_rank": None,
                "state_rank": None,
                "college_code": None,
                "college_name": None,
                "institution_name": None,
                "college_type": None,
                "course": None,
                "category": None,
                "sub_category": None,
                "seat_type": None,
                "quota": None,
                "domicile": None,
                "eligibility": None,
                "score": None,
                "round": None,
            }

            for col_idx, value in enumerate(row):
                field = header_map.get(col_idx)
                if not field:
                    continue
                if field in INT_FIELDS:
                    payload[field] = _to_int(value)
                else:
                    payload[field] = _to_text(value, FIELD_MAX_LEN[field])

            if not payload.get("college_name") and not payload.get("institution_name"):
                continue

            if payload.get("state") is None:
                payload["state"] = state_name

            all_rows.append(payload)
            count += 1

        state_counts[state_name] = count
        ingested_states.append(state_name)

    return all_rows, state_counts, ingested_states


async def replace_cutoff_rows(
    db: AsyncSession,
    rows: List[Dict],
    states: List[str],
) -> int:
    if states:
        await db.execute(
            delete(NeetUg2025Cutoff).where(NeetUg2025Cutoff.state.in_(states))
        )
    if rows:
        await db.execute(insert(NeetUg2025Cutoff), rows)
    await db.flush()
    return len(rows)
